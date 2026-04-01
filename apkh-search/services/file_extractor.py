"""
File content extractor.
Handles PDF, images, docx, xlsx, csv, txt, md files.
Extracts text content for embedding.
"""

import csv
import io
import logging
import os
from pathlib import Path

import fitz  # PyMuPDF
from PIL import Image

logger = logging.getLogger(__name__)

# Try importing pytesseract. OCR is enabled only when the binary is available.
try:
    import pytesseract

    configured_tesseract_cmd = os.getenv("TESSERACT_CMD", "").strip()
    if configured_tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = configured_tesseract_cmd

    # Validate the configured command or PATH installation.
    pytesseract.get_tesseract_version()
    TESSERACT_AVAILABLE = True
except Exception:
    TESSERACT_AVAILABLE = False
    logger.warning(
        "Tesseract binary not found or pytesseract not available. "
        "Set TESSERACT_CMD or add Tesseract to PATH. OCR will be skipped."
    )


def extract_text_from_bytes(file_bytes: bytes, file_name: str) -> dict:
    """
    Extract text from file bytes based on file extension.
    Returns a dict with extracted_text, extraction_method, page_count.
    """
    ext = Path(file_name).suffix.lower()

    try:
        if ext == ".pdf":
            return _extract_pdf(file_bytes, file_name)
        elif ext in (".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tiff"):
            return _extract_image(file_bytes, file_name)
        elif ext in (".txt", ".md"):
            return _extract_text_file(file_bytes, file_name)
        elif ext == ".docx":
            return _extract_docx(file_bytes, file_name)
        elif ext in (".xlsx", ".csv"):
            return _extract_spreadsheet(file_bytes, file_name, ext)
        else:
            logger.warning(f"Unsupported file type: {ext} for file {file_name}")
            return {
                "file_name": file_name,
                "extracted_text": "",
                "extraction_method": "unsupported",
                "page_count": 0,
            }
    except Exception as e:
        logger.error(f"Error extracting text from {file_name}: {e}")
        return {
            "file_name": file_name,
            "extracted_text": "",
            "extraction_method": "error",
            "page_count": 0,
        }


def _extract_pdf(file_bytes: bytes, file_name: str) -> dict:
    """Extract text from PDF. Falls back to OCR if text is too short."""
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    pages_text = []
    extraction_method = "pdf_text"

    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text("text").strip()

        # If text is too short, try OCR on rendered page image.
        if len(text) < 50 and TESSERACT_AVAILABLE:
            try:
                pix = page.get_pixmap(dpi=200)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                ocr_text = pytesseract.image_to_string(img).strip()
                if ocr_text:
                    text = ocr_text
                    extraction_method = "pdf_ocr"
            except Exception as e:
                logger.warning(f"OCR failed on page {page_num + 1} of {file_name}: {e}")

        if text:
            pages_text.append(f"[Page {page_num + 1}]\n{text}")

    page_count = len(doc)
    doc.close()

    return {
        "file_name": file_name,
        "extracted_text": "\n\n".join(pages_text),
        "extraction_method": extraction_method,
        "page_count": page_count,
    }


def _extract_image(file_bytes: bytes, file_name: str) -> dict:
    """Extract text from image using OCR."""
    text = ""
    method = "image_ocr"

    if TESSERACT_AVAILABLE:
        try:
            img = Image.open(io.BytesIO(file_bytes))
            text = pytesseract.image_to_string(img).strip()
        except Exception as e:
            logger.error(f"Image OCR failed for {file_name}: {e}")
            method = "image_ocr_failed"
    else:
        logger.warning(f"Tesseract not available, skipping OCR for {file_name}")
        method = "image_no_tesseract"

    return {
        "file_name": file_name,
        "extracted_text": text,
        "extraction_method": method,
        "page_count": 0,
    }


def _extract_text_file(file_bytes: bytes, file_name: str) -> dict:
    """Read .txt / .md files directly."""
    try:
        text = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1", errors="replace")

    return {
        "file_name": file_name,
        "extracted_text": text.strip(),
        "extraction_method": "direct",
        "page_count": 0,
    }


def _extract_docx(file_bytes: bytes, file_name: str) -> dict:
    """Extract text from .docx files."""
    from docx import Document

    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    text = "\n".join(paragraphs)

    return {
        "file_name": file_name,
        "extracted_text": text,
        "extraction_method": "docx_parse",
        "page_count": 0,
    }


def _extract_spreadsheet(file_bytes: bytes, file_name: str, ext: str) -> dict:
    """Extract text from .xlsx or .csv files."""
    import openpyxl

    text_parts = []

    if ext == ".xlsx":
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                if row_text.strip() and row_text.strip() != "|":
                    text_parts.append(row_text)
        wb.close()
    elif ext == ".csv":
        content = file_bytes.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(content))
        for row in reader:
            row_text = " | ".join(row)
            if row_text.strip():
                text_parts.append(row_text)

    return {
        "file_name": file_name,
        "extracted_text": "\n".join(text_parts),
        "extraction_method": "spreadsheet_parse",
        "page_count": 0,
    }
